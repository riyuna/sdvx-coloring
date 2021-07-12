import requests
import time
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl

songs=dict()
failed=False
try:
    wb=openpyxl.load_workbook('sdvx lv18 table.xlsx')
except:
    print('엑셀 파일을 찾을 수 없습니다.')
    print('같은 폴더 안에 "sdvx lv18 table.xlsx" 파일이 있는지 확인해주세요.')
    time.sleep(10)
    failed=True
if not failed:
    print("안즈인포 아이디를 입력해주세요.")
    id=input()
    print("안즈인포 비밀번호를 입력해주세요. 비밀번호는 수집되지 않습니다.")
    pw=input()
    data={
        "id": id,
        "pw": pw
    }
    table=wb["Sheet1"]
    L=table["C3:L55"]
    checkdict=dict()
    check995dict=dict()
    check998dict=dict()
    checkpucdict=dict()
    def changer(s):
        ss = ''
        if s[0] == '	': s = s[1:]
        for k in s:
            if k == '／':break
            if k == '\n': continue
            if k == '\t':continue
            if k == ' ': continue
            if k == '　': continue
            ss += k.lower()
        for i in range(len(ss)):
            if i>=len(ss):break
            if ss[i]=='&' and ss[i+1:i+5]=='amp;':
                ss=ss[:i+1]+ss[i+5:]
        return ss

    for i in L:
        for j in i:
            s=str(j.value)
            if s==None:continue
            s.strip()
            ss=changer(s)
            if ss=='ελπισ':ss='ελπις'
            songs[ss]=(j.row, j.column)
            checkdict[ss]=False
            check995dict[ss]=False
            check998dict[ss]=False
            checkpucdict[ss]=False

    s='https://anzuinfo.me/login_check.php'
    session=requests.session()

    resp=session.post(s, data=data)
    resp.raise_for_status()
    page=1
    breaked=False
    loginfailed=False
    perm=''
    while True:
        if page>20:break
        url = 'https://anzuinfo.me/myScoreText.html?ver=6&sort=update_down&filter_level=131072&filter_diff=255&filter_comp=31&filter_grade=512&page='
        resp = session.get(url + str(page))
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        text = soup.select('td')
        if text == []:
            loginfailed=True
            break
        for i in text:
            ss=str(i)
            L=ss.split('"')
            if len(L)==1:
                first=L[0]
                first=first[4:len(first)-5]
                for k in first:
                    if k not in '0123456789':continue
                score=int(first)
                if score>=9950000:
                    check995dict[perm]=True
                if score>=9970000:
                    check998dict[perm]=True
                if score==10000000:
                    checkpucdict[perm]=True

            if len(L)>1 and L[1]=='score_title':
                songname=L[2][1:]
                songname=songname[:len(songname)-5]
                songname=songname.lower()
                songname=changer(songname)
                if songname not in songs:
                    continue
                row,col = songs[songname]
                if checkdict[songname]:
                    breaked=True
                    break
                perm=songname
                checkdict[songname]=True

        if breaked:break
        page+=1
    if loginfailed:
        print("로그인에 실패했습니다. 아이디와 비밀번호를 다시 확인해주세요.")
        time.sleep(10)
    else:
        ct=0
        redFill = openpyxl.styles.PatternFill(start_color='FFD3D7',
                           end_color='FFD3D7',
                           fill_type='solid')
        redFill2 = openpyxl.styles.PatternFill(start_color='FFADAD',
                           end_color='FFADAD',
                           fill_type='solid')
        redFill3 = openpyxl.styles.PatternFill(start_color='FF7E7E',
                           end_color='FF7E7E',
                           fill_type='solid')
        yellowFill = openpyxl.styles.PatternFill(start_color='FFFF8E',
                           end_color='FFFF8E',
                           fill_type='solid')
        for i in checkdict:
            if checkdict[i]==False:
                ct+=1
            else:
                row,col=songs[i]
                point=' ABCDEFGHIJKLMN'[col]+str(row)
                table[point].fill=redFill
            if check995dict[i]==True:
                row,col=songs[i]
                point=' ABCDEFGHIJKLMN'[col]+str(row)
                table[point].fill=redFill2
            if check998dict[i]==True:
                row,col=songs[i]
                point=' ABCDEFGHIJKLMN'[col]+str(row)
                table[point].fill=redFill3
            if checkpucdict[i]==True:
                row,col=songs[i]
                point=' ABCDEFGHIJKLMN'[col]+str(row)
                table[point].fill=yellowFill
        wb.save('sdvx lv18 table colored.xlsx')
        print('색칠 완료. 저장된 엑셀파일을 확인해주세요.')
        time.sleep(10)